import openpyxl
import copy
import win32com.client as win32
import pandas as pd
import os
from openpyxl.styles import Font, Border, Side, Alignment 
from openpyxl.utils import get_column_letter
import numpy as np

"""
第一天要改的就是这些文件路径和名字
之后就只需要改today、if_12、ttime
"""
today = '12.26' # 今天的日期
ttime = '9' # 导出的时间
if_12 = 0   # 是否是12点的数据

# 文件名
new_name = r'D:\C 兼辅\D 健康申报\今日未填报清单'+today.replace('.','')+'-'+ttime+'.xlsx' # 最终生成的文件名

benke_name = r'D:\C 兼辅\D 健康申报\本科-今日&连续3日未填报清单'+today.replace('.','')+'.xlsx'
shuobo_name = r'D:\C 兼辅\D 健康申报\硕博-今日&连续3日未填报清单'+today.replace('.','')+'.xlsx'

refer_path = r'D:\C 兼辅\D 健康申报' + '\\'  # 用于匹配的文件路径
refer_name = r'全院学生信息.xlsx'   # 用于匹配的文件名

cal_name = r'D:\C 兼辅\D 健康申报\今日未填报清单1224-9.xlsx'    # 随便一个有班级统计数据的文件
cal_name_12 = r'D:\C 兼辅\D 健康申报\今日未填报清单1224-12.xlsx'

con_path = r'D:\C 兼辅\D 健康申报\健康申报' + '\\'
con_name = r'每日未填报清单.xlsx' # 总表
row_name = r'D:\C 兼辅\D 健康申报\day_row.txt'  # 标记最近三天应该从哪一行开始匹配

data_name = r'C:\Users\97492\Downloads\今日未填报清单.xls'  # 原始下载路径和文件
ttemp_name = r'C:\Users\97492\Downloads\今日未填报清单 (1).xls' 

"""
以下的都不用管了大概也许可能
"""

wb_name = data_name+'x'  # 原地修改格式后的文件名
temp_name = ttemp_name+'x'

# 格式设置
border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))   # 单元格边框
align = Alignment(horizontal='center',vertical='center')    # 单元格居中
data_font = Font(u'宋体',size=15,bold=True)   # 除了“今日未填报清单”的字体
btitle_font = Font(u'宋体',size=20,bold=True)   # “今日未填报清单”的字体
stitle_font = Font(u'宋体',size=10,bold=True)  # 班级数据标题字体加粗
sdata_font = Font(u'宋体',size=10)  # 班级数据除标题行

def style_set(cell_name, cell_type):
    """
    设置字体格式
    输入: 单元格名称, 哪种格式(普通数据:normal_data, "今日未填报清单":btitle, 班级数据标题:stitle, 班级数据:sdata)
    """
    cell_name.alignment = align
    cell_name.border = border
    if cell_type == 'normal_data':
        cell_name.font = data_font
    if cell_type == 'btitle':
        cell_name.font = btitle_font
    if cell_type == 'stitle':
        cell_name.font = stitle_font
    if cell_type == 'sdata':
        cell_name.font = sdata_font

def xls2xlsx(file_name):
    """
    将 xls 文件转换为 xlsx 文件
    输入: 文件名
    """
    excel = win32.Dispatch('Excel.Application')
    data = excel.Workbooks.Open(file_name)
    data.SaveAs(file_name+'x', FileFormat=51)
    data.Close()
    #excel.Application.Quit() 

def func2value(file_name, refer_file, same_file):
    """
    把公式转换为值, 即需要打开一遍文件再保存
    输入: 文件名，公式中引用的文件名，公式中是否没有引用其他文件(引用其他文件为0, 仅引用自己为1)
    """
    excel = win32.Dispatch('Excel.Application')
    if same_file == 0:
        excel.Workbooks.Open(refer_file)   # 这一步必须打开那个被引用的文档，不然就没法转换
    xlApp = win32.Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(file_name)
    xlBook.Save()
    xlBook.Close()
    xlApp.Application.Quit()
    excel.Application.Quit() 

def col_width(sheet_name, twelve, undergraduate):
    """
    调整班级统计数据的列宽（非常简单粗暴的调整方法）
    输入: 要调整的 sheet, 是否是12点的数据, 是否是本科生的数据
    """
    if twelve == 0: # 不是12点的数据
        if undergraduate == 1:  # 本科生的数据
            sheet_name.column_dimensions['A'].width = 22.71
            sheet_name.column_dimensions['B'].width = 30.57
            sheet_name.column_dimensions['C'].width = 30.57
        else:   # 研究生的数据
            sheet_name.column_dimensions['A'].width = 52.43
            sheet_name.column_dimensions['B'].width = 30.57
    else:   # 12点的数据
        if undergraduate == 1:  # 本科生的数据
            sheet_name.column_dimensions['A'].width = 13.57
            sheet_name.column_dimensions['B'].width = 22.29
            sheet_name.column_dimensions['C'].width = 11.14
            sheet_name.column_dimensions['D'].width = 8.43
            sheet_name.column_dimensions['E'].width = 8
            sheet_name.column_dimensions['F'].width = 19.29
        else:   # 研究生的数据
            sheet_name.column_dimensions['A'].width = 52.43
            sheet_name.column_dimensions['B'].width = 15.71
            sheet_name.column_dimensions['C'].width = 19.29

def normal_process(data_file, sdata_file, refer_fpath, refer_file):
    """
    常规处理，即普通的匹配班级后的未填报名单 sheet1
    输入: 除了21级本科生的名单文件, 21级本科生名单文件, 全院学生信息文件路径, 全院学生信息文件
    """

    # 读取文件
    wb = openpyxl.load_workbook(data_file)  # 读取除了21级本科生外的数据
    temp = openpyxl.load_workbook(sdata_file)    # 读取21级本科生的数据
    wb_sheet1 = wb.worksheets[0]    # 数据都在sheet1里
    temp_sheet1 = temp.worksheets[0]

    # 处理21级本科生的名单
    temp_sheet1.unmerge_cells('A1:F1')  # 把“今日未填报清单”拆分
    temp_sheet1.delete_rows(1)  # 删除“今日未填报清单”，Excel的行列是从 1 开始计数的
    temp_sheet1.insert_cols(5)  # 21级本科生少了一列“部门”，先补上

    # 获取需要的表格行列数
    wb_rows = wb_sheet1.max_row # 除了21本科生的行数
    temp_rows = temp_sheet1.max_row # 21本科生的行数
    temp_column = temp_sheet1.max_column # 21本科生的列数 

    # 把21级本科生的数据复制到总表里
    for i in range(2,temp_rows+1):  # 第一行是表头，不需要
        for j in range(1,temp_column+1):
            wb_sheet1.cell(row=wb_rows+i-1,column=j).value = temp_sheet1.cell(row=i,column=j).value
            wb_sheet1.cell(row=wb_rows+i-1,column=j)._style = copy.copy(temp_sheet1.cell(row=i,column=j)._style)    # 保持原来的格式

    # 删除无用的列，包括“身份”（4）、“部门”（5）、“校区”（6）、“是否全日制”（8）
    wb_sheet1.delete_cols(8)
    wb_sheet1.delete_cols(6)
    wb_sheet1.delete_cols(5)
    wb_sheet1.delete_cols(4)

    # 用 VLOOKUP 匹配未填报同学的班级
    wb_rows = wb_sheet1.max_row # 加入了21级本科生的总行数
    wb_sheet1.unmerge_cells('A1:F1')  # 把“今日未填报清单”拆分
    wb_sheet1.delete_rows(1)  # 删除“今日未填报清单”，方便后续操作
    for i in range(2,wb_rows+1):
        #wb_sheet1[f"F{i}"] = f'=VLOOKUP(B{i},\'D:\C 兼辅\D 健康申报\[全院学生信息.xlsx]全院学生信息\'!$B$2:$G$1483,4,0)'
        wb_sheet1[f"F{i}"] = f'=VLOOKUP(B'+str(i)+',\''+refer_fpath+'['+refer_file+']'+refer_name.replace('.xlsx','')+'\'!$B$2:$G$1483,4,0)'
    wb.save(new_name)

    # 把 VLOOKUP 的公式转换为值
    func2value(new_name, refer_path+refer_name, 0)

    # 排序
    wb_data = pd.read_excel(new_name)
    df = pd.DataFrame(wb_data)
    df = df.sort_values('填报状态',ascending=True)
    writer = pd.ExcelWriter(new_name)
    df.to_excel(writer)
    writer.save() 

    # pandas 排序后格式全没了 还得用 openpyxl 处理一下
    wb = openpyxl.load_workbook(new_name,data_only=True)
    wb_sheet1 = wb.worksheets[0] 
    wb_sheet1.delete_cols(1)    # 第一列给它删了，被pd加的序号
    wb_sheet1.insert_rows(1)  # 把那个无用的“今日未填报清单”加回来
    wb_sheet1.cell(row=1,column=1).value = '今日未填报清单'
    wb_sheet1.merge_cells('A1:C1')    # 假惺惺地合并回去
    style_set(wb_sheet1.cell(row=1,column=1),'btitle')  # 修改一下格式

    # 统一格式 重新编号 删除空白的
    wb_rows = wb_sheet1.max_row # 看看现在的行数
    wb_cols = wb_sheet1.max_column # 列数

    na = [] # 需要被删除的 不是我们学院的
    for i in range(3,wb_rows+1):  # 前两行无用
        wb_sheet1.cell(row=i,column=1).value = i-2  # 重新编号
        for j in range(1,wb_cols+1):
            if j==6:
                if wb_sheet1.cell(row=i,column=6).value == None:
                    na.append(i)
            style_set(wb_sheet1.cell(row=i,column=j),'normal_data')  # 修改格式

    na.sort(reverse=True)  
    for num in na:
        wb_sheet1.delete_rows(int(num)) # 删除最后几个 N#A 的

    # 给 sheet1 改个名字，再统一一下列宽
    wb_sheet1.title = '今日未填报清单1' 
    for j in range(1,wb_cols+1):
        style_set(wb_sheet1.cell(row=2,column=j),'normal_data') # 刚刚第二行的格式是没有改的
        wb_sheet1.column_dimensions[get_column_letter(j)].width = 30.57 # 默认是这个就这个吧

    wb.save(new_name)

def class_data(data_file, cal_file, twelve):
    """
    统计各班的数据
    输入: 成品文件, 有班级统计信息的文件（本科生的一个sheet, 研究生的一个sheet）, 是否是12点的文件(12点的文件多了连续三天的)
    """
    wb = openpyxl.load_workbook(data_file, data_only=True)

    # 本科生
    wb.create_sheet()   # 创建一个新的 sheet 放本科生数据

    cal = openpyxl.load_workbook(cal_file)  # 打开参考文件
    wb_sheet2 = wb.worksheets[twelve+1]    # 在 sheet2/3 操作
    cal_sheet2 = cal.worksheets[twelve+1]

    # 获取参考文件的行列
    cal_rows = cal_sheet2.max_row
    cal_column = cal_sheet2.max_column

    # 复制一下本科生的信息 并调整格式
    for i in range(1,cal_rows+1):
        for j in range(1,cal_column+1):
            wb_sheet2.cell(row=i,column=j).value = cal_sheet2.cell(row=i,column=j).value
            style_set(wb_sheet2.cell(row=i,column=j),'sdata')
            if i == 1:
                style_set(wb_sheet2.cell(row=1,column=j),'stitle')
    col_width(wb_sheet2,twelve,1)    # 调整一下列宽

    # 研究生
    wb.create_sheet()   # 创建一个新的 sheet 放研究生数据
    wb_sheet3 = wb.worksheets[twelve+2]    # 在 sheet3/4 操作
    cal_sheet3 = cal.worksheets[twelve+2]

    cal_rows = cal_sheet3.max_row
    cal_column = cal_sheet3.max_column

    for i in range(1,cal_rows+1):
        for j in range(1,cal_column+1):
            wb_sheet3.cell(row=i,column=j).value = cal_sheet3.cell(row=i,column=j).value
            style_set(wb_sheet3.cell(row=i,column=j),'sdata')
            if i == 1:
                style_set(wb_sheet3.cell(row=1,column=j),'stitle')
    col_width(wb_sheet3,twelve,0)    # 调整一下列宽

    wb.save(new_name)

def add_all(data_file, con_fpath, con_file, row_file):
    """
    把12点未填报的数据加入总表
    输入: 12点未填报名单, 总表
    """
    all_data = openpyxl.load_workbook(con_fpath+con_file) # 读取总表
    data_sheet = all_data.worksheets[0]
    data_rows = data_sheet.max_row
    data_sheet.cell(row=data_rows+1,column=1).value = today # 加入今天的日期
    data_sheet.cell(row=data_rows+1,column=1)._style = copy.copy(data_sheet.cell(row=13220,column=1)._style)    # 标黄

    wb = openpyxl.load_workbook(data_file)  # 读取今日未填报清单
    wb_sheet1 = wb.worksheets[0]
    wb_cols =[1,2,3,6]
    wb_rows = wb_sheet1.max_row

    for i in range(3,wb_rows+1): 
        for j in range(4):
            data_sheet.cell(row=data_rows+i-1,column=j+1).value = wb_sheet1.cell(row=i,column=wb_cols[j]).value
            data_sheet.cell(row=data_rows+i-1,column=j+1)._style = copy.copy(data_sheet.cell(row=13542,column=j+1)._style)    # 保持原来的格式
    all_data.save(con_fpath+con_file) 

    row_data = np.loadtxt(row_name)
    row_data[0] = row_data[1]
    row_data[1] = row_data[2]
    row_data[2] = data_rows+2
    np.savetxt(row_name,row_data,fmt="%d")

def add_con3_cell(data_file, con_fpath, con_file, row_file):
    """
    加入连续3天未填报的情况（单元格）
    输入: 未填报名单，总表
    """
    wb = openpyxl.load_workbook(data_file)  # 读取今日未填报清单
    wb_sheet1 = wb.worksheets[0]
    wb_sheet1.cell(row=2,column=7).value = '近3日未填报次数'
    style_set(wb_sheet1.cell(row=2,column=7),'normal_data')
    wb_sheet1.column_dimensions[get_column_letter(7)].width = 30.57
    wb_rows = wb_sheet1.max_row

    all_data = openpyxl.load_workbook(con_fpath+con_file) # 读取总表
    data_sheet = all_data.worksheets[0]
    data_rows = data_sheet.max_row
    #print(data_rows)

    row_data = np.loadtxt(row_file)
    print(int(row_data[0]))

    for i in range(3,wb_rows+1):
        #wb_sheet1[f"G{i}"] = f'=COUNTIF(\'D:\C 兼辅\D 健康申报\健康申报\[每日未填报清单.xlsx]sheet\'!$B$'+str(qiantian)+':$B$'+str(data_rows)+',B'+str(i)+')'
        wb_sheet1[f"G{i}"] = f'=COUNTIF(\''+con_fpath+'['+con_file+']sheet\'!$B$'+str(int(row_data[0]))+':$B$'+str(data_rows)+',B'+str(i)+')'
    wb.save(new_name)

    func2value(data_file, con_fpath+con_file, 0)

def add_con3_sheet(data_file):
    """
    创建连续3天未填报的sheet
    输入: 未填报名单
    """
    wb = openpyxl.load_workbook(data_file, data_only=True) 
    wb_sheet1 = wb.worksheets[0]
    for i in range(3, wb_sheet1.max_row+1):
        style_set(wb_sheet1.cell(row=i, column=7),'normal_data')
    wb.create_sheet()   # 创建新sheet
    wb_sheett = wb.worksheets[1]
    wb_sheett.title = '今日未填报清单2'
    wb_cols = wb_sheet1.max_column
    wb_rows = wb_sheet1.max_row

    count = 3

    for i in range(1,3): 
        for j in range(1,wb_cols+1):
            wb_sheett.cell(row=i,column=j).value = wb_sheet1.cell(row=i,column=j).value
            wb_sheett.cell(row=i,column=j)._style = copy.copy(wb_sheet1.cell(row=i,column=j)._style)

    for i in range(3,wb_rows+1): 
        if wb_sheet1.cell(row=i,column=7).value==3:
            for j in range(1,wb_cols+1):
                wb_sheett.cell(row=count,column=j).value = wb_sheet1.cell(row=i,column=j).value
                wb_sheett.cell(row=count,column=j)._style = copy.copy(wb_sheet1.cell(row=i,column=j)._style)
            style_set(wb_sheett.cell(row=count,column=7),'normal_data')
            count += 1

    for j in range(1,wb_sheett.max_column+1):
        wb_sheett.column_dimensions[get_column_letter(j)].width = 30.57 # 调整列宽

    wb2_rows = wb_sheett.max_row # 看看现在的行数

    for i in range(3,wb2_rows+1):  # 前两行无用
        wb_sheett.cell(row=i,column=1).value = i-2  # 重新编号
        style_set(wb_sheett.cell(row=i,column=j),'normal_data')

    wb_sheett.merge_cells('A1:C1')    # 假惺惺地合并回去

    wb.save(new_name)

def del_file():
    """
    删除下载的 2 个 xls 和生成的 2 个 xlsx 文件
    """
    file_name_list = [data_name, ttemp_name, wb_name, temp_name]
    for fname in file_name_list:
        if os.path.exists(fname):
            os.remove(fname) 

def benke(data_file, benke_file):
    """
    生成最后的本科表格
    输入: 今日未填报清单, 最后要输出的表格
    """
    wb = openpyxl.load_workbook(data_file)
    wb_sheet1 = wb.worksheets[0]
    wb_rows1 = wb_sheet1.max_row
    del1 = []

    for i in range(3,wb_rows1+1):
        if wb_sheet1.cell(row=i, column=4).value == '博士' or wb_sheet1.cell(row=i, column=4).value == '硕士':
            del1.append(i)
    del1.sort(reverse=True)  
    for num in del1:
        wb_sheet1.delete_rows(int(num))
    for i in range(3,wb_sheet1.max_row+1):  # 前两行无用
        wb_sheet1.cell(row=i,column=1).value = i-2  # 重新编号
    
    wb_sheet2 = wb.worksheets[1]
    wb_rows2 = wb_sheet2.max_row
    del2 = []

    for i in range(3,wb_rows2+1):
        if wb_sheet2.cell(row=i, column=4).value == '博士' or wb_sheet1.cell(row=i, column=4).value == '硕士':
            del2.append(i)
    for num in del2:
        wb_sheet2.delete_rows(int(num))       
    for i in range(3,wb_sheet2.max_row):  # 前两行无用
        wb_sheet2.cell(row=i,column=1).value = i-2  # 重新编号
    
    wb.remove(wb.worksheets[3])

    wb.save(benke_file)

def shuobo(data_file, shuobo_file):
    """
    生成最后的硕博表格
    输入: 今日未填报清单, 最后要输出的表格
    """
    wb = openpyxl.load_workbook(data_file)
    wb_sheet1 = wb.worksheets[0]
    wb_rows1 = wb_sheet1.max_row
    del1 = []

    for i in range(3,wb_rows1+1):
        if wb_sheet1.cell(row=i, column=4).value == '本科生':
            del1.append(i)
    del1.sort(reverse=True)  
    for num in del1:
        wb_sheet1.delete_rows(int(num))
    for i in range(3,wb_sheet1.max_row+1):  # 前两行无用
        wb_sheet1.cell(row=i,column=1).value = i-2  # 重新编号
    
    wb_sheet2 = wb.worksheets[1]
    wb_rows2 = wb_sheet2.max_row
    del2 = []

    for i in range(3,wb_rows2+1):
        if wb_sheet2.cell(row=i, column=4).value == '本科生':
            del2.append(i)
    del2.sort(reverse=True)
    for num in del2:
        wb_sheet2.delete_rows(int(num))       
    for i in range(3,wb_sheet2.max_row):  # 前两行无用
        wb_sheet2.cell(row=i,column=1).value = i-2  # 重新编号
    
    wb.remove(wb.worksheets[2])

    wb.save(shuobo_file)


"""
虽然但是 我不想写主函数
"""
# 把下载的 2 个文件转换成 xlsx
xls2xlsx(data_name)
xls2xlsx(ttemp_name)

# 正常地处理匹配班级数据
normal_process(wb_name, temp_name, refer_path, refer_name)

# 加入班级数据统计
#if_12 = 0   # 是否是12点的数据
if if_12 == 0:  # 非12点
    class_data(new_name, cal_name, if_12) 
else:   # 12点
    add_all(new_name,con_path, con_name, row_name)
    add_con3_cell(new_name, con_path, con_name, row_name)
    add_con3_sheet(new_name)
    class_data(new_name, cal_name_12, if_12) 
    benke(new_name, benke_name)
    shuobo(new_name,shuobo_name)

# 删除下载的文件
del_file()
